import openpyxl,requests,schedule,gspread,re,time
from openpyxl.utils import get_column_letter
from datetime import datetime,timedelta
from oauth2client.service_account import ServiceAccountCredentials
from google.oauth2 import service_account
from collections import defaultdict
import pandas as pd
import os
from dotenv import load_dotenv
import json
import io

load_dotenv()


def parsing():
    # url на аналитику
    url = 'https://seller-analytics-api.wildberries.ru/api/v2/nm-report/detail'
    # url на получение id
    url_1 = 'https://advert-api.wb.ru/adv/v1/promotion/count'
    # url на получение статистики
    url_2 = 'https://advert-api.wb.ru/adv/v2/fullstats'

    API_KEY = os.getenv("API_KEY")
    NUTRA = os.getenv('SECRET_JSON')
    KEY_TABLE = os.getenv('KEY_TABLE')
    #print(KEY_TABLE)


    HeaderApiKey1 = {
        'Authorization': f'{API_KEY}',
        'Content-Type': 'application/json'
    }
    print(HeaderApiKey1)

    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    credentials = ServiceAccountCredentials.from_json_keyfile_name("secret.json", scope)

    columnStat = 3

    now = datetime.now()
    now = now - timedelta(days=1)
    start_of_month = datetime(now.year, now.month, 1)
    dates = pd.date_range(start_of_month, now, freq='D')
    newdates = dates
    saved_positions = {}
    for date in newdates:
        next_day = False
        Jdata = None

        Jdata1 = None
        while next_day == False:
            data = {
                "brandNames": [],
                "timezone": "Europe/Moscow",
                "period": {
                    "begin": date.strftime("%Y-%m-%d %H:%M:%S"),
                    "end": (date + timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S")
                },
                "orderBy": {
                    "field": "ordersSumRub",
                    "mode": "desc"
                },
                "page": 1
            }
            time.sleep(1)
            print(date)
            print()
            response = requests.post(url, json=data, headers=HeaderApiKey1)
            response1 = requests.get(url_1, headers=HeaderApiKey1)

            if response.status_code == 200:
                print('Данные успешно получены за', date.strftime("%Y-%m-%d"))
                print()
                Jdata = response.json()
                Jdata1 = response1.json()
                next_day = True
            else:
                if response.status_code != 200:
                    print(response.status_code)
                    Jdata = Jdata
                    Jdata1 = Jdata1
                    next_day = False
                    continue

            wb = openpyxl.Workbook()
            sheet = wb.active

            def deleate(res):
                while re.search('previousPeriod.*?stocks', str(res), flags=re.DOTALL):
                    res = re.sub('previousPeriod.*?stocks', '', str(res), flags=re.DOTALL)
                return res

            def Remove(arr):
                cleaned_arr = []
                for word in arr:
                    word = ''.join(e for e in word if e.isalnum() or e.isspace())
                    if word:
                        cleaned_arr.append(word)
                return cleaned_arr

            def SpaceX(cleaned_arr):
                return ''.join(cleaned_arr)

            pop = deleate(Jdata)
            pop1 = deleate(Jdata1)

            cleaned_text = Remove(str(pop))
            cleaned_text1 = Remove(str(pop1))

            result = SpaceX(cleaned_text)
            result1 = SpaceX(cleaned_text1)

            column = 2

            words = result.split()
            words1 = result1.split()

            start_row = 2
            start_column = 3

            start_of_month = datetime(now.year, now.month, 1)
            date_range = pd.date_range(start_of_month, now, freq='d')

            indices_id = [i for i, x in enumerate(words1) if x == "advertId"]
            id = [int(words1[i + 1]) if i + 1 < len(words1) else None for i in indices_id]
            date_from = date.strftime("%Y-%m-%d")
            next_day2 = False

            while next_day2 == False:
                params1 = [{'id': c, 'dates': [date_from]} for c in id]
                response2 = requests.post(url_2, headers=HeaderApiKey1, json=params1)
                if response2.status_code == 200:

                    next_day2 = True
                else:
                    if response2.status_code != 200:
                        print(response2.json())
                        print('loading2.....')
                        next_day2 = False
                        continue

            Jdata2 = response2.json()
            camp_data = []
            for c in Jdata2:
                for d in c['days']:
                    for a in d['apps']:
                        for nm in a['nm']:
                            nm['appType'] = a['appType']
                            nm['date'] = d['date']
                            nm['advertId'] = c['advertId'] 
                            camp_data.append(nm)
            camp_df = pd.DataFrame(camp_data)
            df_filtered = camp_df[["nmId", 'views', "clicks", "advertId"]]
            df_filtered = camp_df.groupby('advertId').agg(
                {'nmId': 'first', 'views': 'sum', 'clicks': 'sum'}).reset_index()
            df_filtered = df_filtered.groupby('nmId').agg(
                lambda x: x.sum() if x.name != 'advertId' else x.iloc[0]).reset_index() 
            df_filtered.drop(columns=['advertId'], inplace=True)
            df_filtered['CTR'] = (round(df_filtered['clicks'] / df_filtered['views'] * 100, 2))
            camp_data1 = df_filtered.set_index('nmId').to_dict(orient="index")

            for k, v in camp_data1.items():
                camp_data1[k]['Показы'] = v.pop('views')
                camp_data1[k]['Клики'] = v.pop('clicks')
                camp_data1[k]['CTR'] = v.pop('CTR')

            found_brand = False
            buffer = []
            brands = []

            for word in words:
                if word == "brandName":
                    found_brand = True
                    if buffer:
                        brand_name = ' '.join(buffer)
                        brand_name = brand_name.replace("brandName", "")
                        brands.append(brand_name)
                        buffer = []
                elif word == "object":
                    found_brand = False

                if found_brand:
                    buffer.append(word)

            if buffer:
                brand_name = ' '.join(buffer)
                brand_name = brand_name.replace("brandName", "")
                brands.append(brand_name)

            indices_name = [i for i, x in enumerate(words) if x == "name"]
            name = [words[i + 1] if i + 1 < len(words) else None for i in indices_name]



            indices_nmID = [i for i, x in enumerate(words) if x == "nmID"]
            nmID = [words[i + 1] if i + 1 < len(words) else None for i in indices_nmID]
            indices_ost = [i for i, x in enumerate(words) if x == "stocksWb"]
            stocksWb = [words[i + 1] if i + 1 < len(words) else None for i in indices_ost]
            indices_o = [i for i, x in enumerate(words) if x == "openCardCount"]
            openCardCount = [words[i + 1] if i + 1 < len(words) else None for i in indices_o]
            indices_a = [i for i, x in enumerate(words) if x == "addToCartPercent"]
            addToCartPercent = [words[i + 1] if i + 1 < len(words) else None for i in indices_a]
            indices_c = [i for i, x in enumerate(words) if x == "cartToOrderPercent"]
            cartToOrderPercent = [words[i + 1] if i + 1 < len(words) else None for i in indices_c]
            indices_aa = [i for i, x in enumerate(words) if x == "addToCartCount"]
            addToCartCount = [words[i + 1] if i + 1 < len(words) else None for i in indices_aa]
            combined_list = []


            max_len = max(len(brands), len(openCardCount), len(addToCartPercent), len(cartToOrderPercent),
                          len(addToCartCount), len(stocksWb), len(nmID), len(name))











            for i in range(max_len):
                if i < len(name):
                    combined_list.append(name[i])
                if i < len(brands):
                    combined_list.append("brand: " + brands[i])
                if i < len(openCardCount):
                    combined_list.append("Переходы: " + openCardCount[i])
                if i < len(addToCartPercent):
                    combined_list.append("Конверсии в корзину: " + addToCartPercent[i])
                if i < len(cartToOrderPercent):
                    combined_list.append("Конверсии в заказ: " + cartToOrderPercent[i])
                if i < len(addToCartCount):
                    combined_list.append("Добавление в корзину: " + addToCartCount[i])
                if i < len(stocksWb):
                    combined_list.append("Остатки товаров на складе: " + stocksWb[i])
                if i < len(nmID):
                    combined_list.append("ID: " + nmID[i])



            brand_data = defaultdict(
                lambda: {'Переходы': 0, 'Конверсии в корзину': 0, 'Конверсии в заказ': 0, 'Добавление в корзину': 0,
                         'Остатки товаров на складе': 0, 'Бренд': "", 'Показы': '-', 'Клики': "-", 'CTR': '-'})
            IDD = []

            idol = 0
            while idol < len(combined_list):
                if combined_list[idol].startswith('Возбуждающие'):
                    del combined_list[idol:idol + 8]
                elif combined_list[idol].startswith('Лубриканты'):
                    del combined_list[idol:idol + 8]
                else:
                    idol += 1



            #new_list = [item for item in combined_list if 'Жиросжигатели' not in item]



            print(combined_list)
            print(len(combined_list))


            for i in range(0, len(combined_list), 8):
                brand = combined_list[i + 1].split(': ')[1]
                openCardCount = int(combined_list[i + 2].split(': ')[1])
                addToCartPercent = int(combined_list[i + 3].split(': ')[1])
                cartToOrderPercent = int(combined_list[i + 4].split(': ')[1])
                addToCartCount = int(combined_list[i + 5].split(': ')[1])
                stocksWb = int(combined_list[i + 6].split(': ')[1])
                nmID1 = int(combined_list[i + 7].split(': ')[1])

                brand_data[nmID1]['Переходы'] = openCardCount
                brand_data[nmID1]['Конверсии в корзину'] = addToCartPercent
                brand_data[nmID1]['Конверсии в заказ'] = cartToOrderPercent
                brand_data[nmID1]['Добавление в корзину'] = addToCartCount
                brand_data[nmID1]['Остатки товаров на складе'] = stocksWb
                brand_data[nmID1]['Бренд'] = brand
                brand_data[nmID1]['ID'] = nmID1
                IDD.append(nmID1)

            for key in camp_data1.keys():
                if key in brand_data.keys():
                    brand_data[key].update(camp_data1[key])
                    brand_data[key]['Показы'] = camp_data1[key]['Показы']
                    brand_data[key]['Клики'] = camp_data1[key]['Клики']
                    brand_data[key]['CTR'] = camp_data1[key]['CTR']

            row_ro = 3
            row_pi = 2

            for ID in IDD:

                sheet.cell(row=row_pi, column=column, value="НАЗВАНИЕ БРЕНДА")
                sheet.cell(row=row_pi, column=column - 1, value="МЕТРИКА")
                sheet.cell(row=row_pi, column=column + 1, value="Артикул")

                row_pi += 10
                if ID in brand_data:
                    data = brand_data[ID]

                    if ID in saved_positions:
                        row_ro = saved_positions[ID]
                    else:
                        row_ro += 1
                        saved_positions[ID] = row_ro

                    sheet.cell(row=row_ro, column=column - 1, value="Переходы")
                    sheet.cell(row=row_ro, column=columnStat + 1, value=data['Переходы'])
                    sheet.cell(row=row_ro, column=column, value=data['Бренд'])
                    sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
                    row_ro += 1
                    sheet.cell(row=row_ro, column=column - 1, value="Конверсии в корзину")
                    sheet.cell(row=row_ro, column=columnStat + 1, value=str(data['Конверсии в корзину']) + "%")
                    sheet.cell(row=row_ro, column=column, value=data['Бренд'])
                    sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
                    row_ro += 1
                    sheet.cell(row=row_ro, column=column - 1, value="Конверсии в заказ")
                    sheet.cell(row=row_ro, column=columnStat + 1, value=str(data['Конверсии в заказ']) + "%")
                    sheet.cell(row=row_ro, column=column, value=data['Бренд'])
                    sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
                    row_ro += 1
                    sheet.cell(row=row_ro, column=column - 1, value="Добавление в корзину")
                    sheet.cell(row=row_ro, column=columnStat + 1, value=data['Добавление в корзину'])
                    sheet.cell(row=row_ro, column=column, value=data['Бренд'])
                    sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
                    row_ro += 1
                    sheet.cell(row=row_ro, column=column - 1, value="Остатки товаров на складе")
                    sheet.cell(row=row_ro, column=columnStat + 1, value=data['Остатки товаров на складе'])
                    sheet.cell(row=row_ro, column=column, value=data['Бренд'])
                    sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
                    row_ro += 1

                    sheet.cell(row=row_ro, column=column - 1, value="Показы")
                    sheet.cell(row=row_ro, column=columnStat + 1, value=data['Показы'])
                    sheet.cell(row=row_ro, column=column, value=data['Бренд'])
                    sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
                    row_ro += 1
                    sheet.cell(row=row_ro, column=column - 1, value="Клики")
                    sheet.cell(row=row_ro, column=columnStat + 1, value=data['Клики'])
                    sheet.cell(row=row_ro, column=column, value=data['Бренд'])
                    sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
                    row_ro += 1
                    sheet.cell(row=row_ro, column=column - 1, value="CTR")
                    sheet.cell(row=row_ro, column=columnStat + 1, value=data['CTR'])
                    sheet.cell(row=row_ro, column=column, value=data['Бренд'])

                    sheet.cell(row=row_ro, column=column + 1, value=data['ID'])

                    row_ro += 3
                    row_pi += 1

                for i, date in enumerate(date_range):
                    sheet.cell(row=start_row, column=start_column + i + 1, value=date.strftime('%d.%m.%y'))
                    row = start_row
                start_row += 11
            wb.save("analyticWB.xlsx")


            def CopyFromExcInGsh(): 
                client = gspread.authorize(credentials)

                spreadsheet = client.open(KEY_TABLE)
                worksheet = spreadsheet.worksheet('Аналитика и статистика космодом')

                df = pd.read_excel("analyticWB.xlsx")
                data_list = df.values.tolist()
                num_cols = len(data_list[0])

                cell_list = worksheet.range('A1:' + get_column_letter(num_cols) + str(len(data_list)))
                for cell in cell_list:
                    row = (cell.row - 1) if (cell.row - 1) < len(data_list) else -1
                    col = (cell.col - 1) if (cell.col - 1) < num_cols else -1
                    if row != -1 and col != -1:
                        value = data_list[row][col]
                        if pd.notna(value):
                            cell.value = str(value)

                worksheet.update_cells(cell_list)
                print("Данные успешно загружены в таблицу Google Sheets!")

            if response.status_code == 200:
                CopyFromExcInGsh()
                columnStat += 1
        
parsing()